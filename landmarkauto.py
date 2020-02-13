import os
from openpyxl import load_workbook
import cv2
import dlib
import xlwt
import xlwings
import matplotlib.pyplot as plt
from matplotlib import style
style.use("ggplot")
from tempfile import TemporaryFile

directory = 'AbuBakarData\Happy'

src_wb = load_workbook('Book2.xlsx')
wb = xlwings.Book("D:\d.xlsx")
Sheet1 = wb.sheets[8]
Sheet2 = wb.sheets[7]


sheets = src_wb.sheetnames
Sheet1 = src_wb[sheets[0]]

dest_wb = load_workbook('landmarks.xlsx')
src_sheet = src_wb.get_sheet_by_name('Sheet1')
dest_sheet = dest_wb.get_sheet_by_name('Sheet1')
j=1
for filename in os.listdir(directory):
    if filename.endswith(".jpg") or filename.endswith(".py"):
        images=os.path.join(directory, filename)
        image = cv2.imread(images)
        # Set up some required objects
        detector = dlib.get_frontal_face_detector()  # Face detector
        predictor = dlib.shape_predictor(
            "shape_predictor_68_face_landmarks.dat")  # Landmark identifier. Set the filename to whatever you named the downloaded file


        def get_landmarks(image):
            xlist = []
            ylist = []
            landmarks = []
            detections = detector(image, 1)
            for k, d in enumerate(detections):  # For all detected face instances individually
                shape = predictor(image, d)  # Draw Facial Landmarks with the predictor class

                for i in range(1, 68):  # Store X and Y coordinates in two lists
                    xlist.append(float(shape.part(i).x))
                    ylist.append(float(shape.part(i).y))
                for x, y in zip(xlist, ylist):  # Store all landmarks in one list in the format x1,y1,x2,y2,etc.
                    landmarks.append(x)
                    landmarks.append(y)

            if len(detections) > 0:
                return landmarks

            else:  # If no faces are detected, return error message to other function to handle
                landmarks = "error"
                return landmarks


        list = []
        list = get_landmarks(image)
        print(list)
        for i, e in enumerate(list):
            print(e)
            print(j)
            print(i)
            Sheet1.cell(row=i, column=j).value = e  # This will change the cell(2,4) to 4

        src_wb.save("D:\d.xlsx")
        j = j + 1

        for i in range(1, src_sheet.max_row + 1):
            for j in range(1, src_sheet.max_column + 1):
                dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

        src_wb.save('Book2.xlsx')
        dest_wb.save('landmarks.xlsx')

    else:
        continue