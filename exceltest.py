from openpyxl import load_workbook

src_wb = load_workbook('Book1.xlsx')
dest_wb = load_workbook('Book2.xlsx')

src_sheet = src_wb.get_sheet_by_name('Sheet1')
dest_sheet = dest_wb.get_sheet_by_name('Sheet1')

for i in range(1, src_sheet.max_row+1):
    for j in range(1, src_sheet.max_column+1):
        dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

src_wb.save('Book1.xlsx')
dest_wb.save('Book2.xlsx')
