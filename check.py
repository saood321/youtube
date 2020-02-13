import os
from copy import copy

from openpyxl import load_workbook
import cv2
import dlib
import xlwt
import xlwings
import matplotlib.pyplot as plt
from matplotlib import style
style.use("ggplot")
from tempfile import TemporaryFile

directory = 'AbuBakarData\Happy'
src_wb = xlwings.Book("D:\d.xlsx")
Sheet1 = src_wb.sheets[0]
print(Sheet1)



j=1
for filename in os.listdir(directory):
    if filename.endswith(".jpg") or filename.endswith(".py"):
        images=os.path.join(directory, filename)
        image = cv2.imread(images)
        # Set up some required objects
        detector = dlib.get_frontal_face_detector()  # Face detector
        predictor = dlib.shape_predictor(
            "shape_predictor_68_face_landmarks.dat")  # Landmark identifier. Set the filename to whatever you named the downloaded file


        def get_landmarks(image):
            xlist = []
            ylist = []
            landmarks = []
            detections = detector(image, 1)
            for k, d in enumerate(detections):  # For all detected face instances individually
                shape = predictor(image, d)  # Draw Facial Landmarks with the predictor class

                for i in range(1, 68):  # Store X and Y coordinates in two lists
                    xlist.append(float(shape.part(i).x))
                    ylist.append(float(shape.part(i).y))
                for x, y in zip(xlist, ylist):  # Store all landmarks in one list in the format x1,y1,x2,y2,etc.
                    landmarks.append(x)
                    landmarks.append(y)

            if len(detections) > 0:
                return landmarks

            else:  # If no faces are detected, return error message to other function to handle
                landmarks = "error"
                return landmarks


        list = []
        list = get_landmarks(image)
        print(list)
        for i, e in enumerate(list):
            Sheet1.range(i, j).value = e
        src_wb.save()
        src_wb.close()
        j = j + 1

        for i in range(1, src_sheet.max_row + 1):
            for j in range(1, src_sheet.max_column + 1):
                dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

        src_wb.save('Book2.xlsx')
        dest_wb.save('landmarks.xlsx')

    else:
        continue